import os
import sys
import json
import urllib.request
import urllib.error
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_data(ip: str) -> dict:
    url_info = f"http://{ip}:8001/info"
    url_metrics = f"http://{ip}:8001/metrics"
    
    try:
        req_info = urllib.request.Request(url_info)
        with urllib.request.urlopen(req_info, timeout=5) as response:
            info = json.loads(response.read())
            
        req_metrics = urllib.request.Request(url_metrics)
        with urllib.request.urlopen(req_metrics, timeout=5) as response:
            metrics = json.loads(response.read())
            
        return {"info": info, "metrics": metrics}
    except urllib.error.URLError as e:
        print(f"\n[X] Error conectando a la IP {ip}: {e}")
        print("Asegurate de que la PC esté encendida, el agente instalado, y la IP sea correcta.")
        sys.exit(1)
    except Exception as e:
        print(f"\n[X] Error inesperado: {e}")
        sys.exit(1)

def style_header(cell, ws, row, col_start, col_end):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    b = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = b

def format_gb(val):
    if val is None: return "—"
    return f"{val} GB"

def format_pct(val):
    if val is None: return "—"
    return f"{val}%"

def generate_excel(ip: str, data: dict):
    info = data["info"]
    metrics = data["metrics"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte PC"
    
    # Titulo Principal
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Reporte de Monitoreo - {info.get('hostname', ip)}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A2'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, color="555555")
    
    current_row = 4
    
    def add_section(title, data_dict):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell = ws.cell(row=current_row, column=1, value=title.upper())
        style_header(cell, ws, current_row, 1, 2)
        current_row += 1
        
        for key, value in data_dict.items():
            ws.cell(row=current_row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=str(value))
            # Borders
            thin = Side(border_style="thin", color="CCCCCC")
            ws.cell(row=current_row, column=1).border = Border(bottom=thin, right=thin)
            ws.cell(row=current_row, column=2).border = Border(bottom=thin)
            current_row += 1
        current_row += 1

    # Seccion 1: Sistema y Red
    add_section("Sistema Operativo", {
        "Hostname": info.get("hostname"),
        "IP Address": info.get("ip", ip),
        "Sistema Operativo": info.get("os"),
        "Edición": info.get("os_edition"),
        "Arquitectura": info.get("architecture"),
    })
    
    add_section("Hardware", {
        "Procesador": info.get("processor"),
        "Núcleos Físicos": metrics.get("cpu", {}).get("physical_cores"),
        "Núcleos Lógicos": metrics.get("cpu", {}).get("logical_cores"),
        "RAM Total": format_gb(metrics.get("memory", {}).get("total_gb")),
        "RAM En Uso": format_gb(metrics.get("memory", {}).get("used_gb")),
        "RAM %": format_pct(metrics.get("memory", {}).get("percent")),
        "Swap Total": format_gb(metrics.get("swap", {}).get("total_gb")),
        "Batería": f"{metrics.get('battery', {}).get('percent', '—')}%" if metrics.get('battery') else "No detectada",
    })
    
    # Discos
    disks = metrics.get("disk", {})
    if disks:
        disk_data = {}
        for dev, d in disks.items():
            disk_data[dev] = f"{d.get('percent', 0)}% usado ({d.get('used_gb')} GB / {d.get('total_gb')} GB)"
        add_section("Discos Duros", disk_data)
        
    # Red
    net = metrics.get("network", {})
    if net:
        add_section("Tráfico de Red", {
            "Conexiones Activas": net.get("connections", "—"),
            "Gigabytes Descargados": format_gb(round(net.get("bytes_recv", 0) / (1024**3), 2)),
            "Gigabytes Subidos": format_gb(round(net.get("bytes_sent", 0) / (1024**3), 2)),
        })
        
    # Usuarios
    users = metrics.get("users", [])
    if users:
        user_list = [u.get("name", "Desconocido") for u in users]
        add_section("Usuarios Logueados", {
            "Usuarios": ", ".join(user_list)
        })
        
    # Top 5 Procesos
    processes = metrics.get("processes", {}).get("top_cpu", [])[:5]
    if processes:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=1, value="TOP 5 PROCESOS QUE MÁS CONSUMEN")
        style_header(cell, ws, current_row, 1, 4)
        current_row += 1
        
        headers = ["Nombre", "PID", "CPU %", "RAM (MB)"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=current_row, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E0E0E0")
        current_row += 1
        
        for p in processes:
            ws.cell(row=current_row, column=1, value=p.get('name'))
            ws.cell(row=current_row, column=2, value=p.get('pid'))
            ws.cell(row=current_row, column=3, value=f"{p.get('cpu_percent', 0)}%")
            ws.cell(row=current_row, column=4, value=p.get('memory_mb'))
            current_row += 1

    # Auto ajustar anchos
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    # Guardar
    os.makedirs("Reportes", exist_ok=True)
    filename = f"Reportes/Reporte_IP_{ip.replace('.', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    print("=" * 60)
    print("   GENERADOR DE REPORTES EXCEL - CODERE PC MONITOR")
    print("=" * 60)
    print("Este programa se conectará al agente de la computadora deseada")
    print("y extraerá absolutamente todos los datos a un archivo Excel.")
    print("-" * 60)
    
    if len(sys.argv) > 1:
        ip_target = sys.argv[1]
    else:
        ip_target = input("\n> Ingrese la IP de la computadora: ").strip()
        
    if not ip_target:
        print("[X] IP no válida.")
        sys.exit(1)
        
    print(f"\n[~] Conectando a {ip_target}...")
    data = get_data(ip_target)
    
    print("[~] Extracción exitosa. Generando Excel...")
    output_file = generate_excel(ip_target, data)
    
    print(f"\n[OK] ¡EXCEL GENERADO CON ÉXITO!")
    print(f"Archivo guardado en: {os.path.abspath(output_file)}")
    
    # Intentar abrir la carpeta en Windows
    try:
        if os.name == 'nt':
            os.startfile(os.path.abspath("Reportes"))
    except:
        pass
    
    if len(sys.argv) == 1:
        print("\nPresione ENTER para salir.")
        input()
