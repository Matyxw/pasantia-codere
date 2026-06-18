#!/usr/bin/env python3
"""
CLIENTE CENTRAL - Panel de control para monitorear múltiples PCs
Corre en la máquina "núcleo" que quiere monitorear las otras
"""

import requests
import json
import time
from datetime import datetime
from pathlib import Path
import os

class ClienteCentral:
    def __init__(self, archivo_ips="ips_registradas.json"):
        self.archivo_ips = archivo_ips
        self.ips_registradas = self.cargar_ips()
        self.timeout = 5
    
    def cargar_ips(self):
        """Carga las IPs registradas desde archivo"""
        if os.path.exists(self.archivo_ips):
            try:
                with open(self.archivo_ips, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def guardar_ips(self):
        """Guarda las IPs registradas"""
        with open(self.archivo_ips, 'w') as f:
            json.dump(self.ips_registradas, f, indent=2)
    
    def registrar_ip(self, ip, nombre_pc):
        """Registra una nueva IP"""
        self.ips_registradas[ip] = {
            "nombre": nombre_pc,
            "fecha_registro": datetime.now().isoformat(),
            "estado": "desconocido"
        }
        self.guardar_ips()
        print(f"✅ {nombre_pc} ({ip}) registrada")
    
    def verificar_conexion(self, ip):
        """Verifica si una PC está conectada"""
        try:
            respuesta = requests.get(f"http://{ip}:5000/salud", timeout=self.timeout)
            return respuesta.status_code == 200
        except:
            return False
    
    def obtener_datos(self, ip):
        """Obtiene todos los datos de una PC específica"""
        try:
            respuesta = requests.get(f"http://{ip}:5000/datos", timeout=self.timeout)
            if respuesta.status_code == 200:
                datos = respuesta.json()
                datos["timestamp"] = datetime.now().isoformat()
                return datos
            return None
        except Exception as e:
            return {"error": f"No se pudo conectar: {str(e)}"}
    
    def ejecutar_comando(self, ip, comando):
        """Ejecuta un comando en una PC remota"""
        try:
            respuesta = requests.post(
                f"http://{ip}:5000/ejecutar",
                json={"comando": comando},
                timeout=self.timeout
            )
            return respuesta.json()
        except Exception as e:
            return {"error": str(e)}
    
    def mostrar_menu(self):
        """Muestra el menú principal"""
        while True:
            print("\n" + "="*50)
            print("🖥️  PANEL CENTRAL DE MONITOREO")
            print("="*50)
            print("1. Registrar nueva PC")
            print("2. Ver todas las PCs")
            print("3. Ver datos de una PC específica")
            print("4. Ejecutar comando remoto")
            print("5. Exportar a Excel")
            print("6. Verificar estado de todas")
            print("7. Salir")
            print("="*50)
            
            opcion = input("Selecciona opción (1-7): ").strip()
            
            if opcion == "1":
                self.registrar_nueva_pc()
            elif opcion == "2":
                self.ver_todas_pcs()
            elif opcion == "3":
                self.ver_datos_pc()
            elif opcion == "4":
                self.ejecutar_comando_remoto()
            elif opcion == "5":
                self.exportar_excel()
            elif opcion == "6":
                self.verificar_estado_todas()
            elif opcion == "7":
                print("👋 Hasta luego!")
                break
            else:
                print("❌ Opción inválida")
    
    def registrar_nueva_pc(self):
        """Registra una nueva PC"""
        ip = input("Ingresa la IP de la PC: ").strip()
        nombre = input("Ingresa un nombre para la PC: ").strip()
        
        if self.verificar_conexion(ip):
            self.registrar_ip(ip, nombre)
        else:
            print(f"❌ No se puede conectar a {ip}")
    
    def ver_todas_pcs(self):
        """Muestra todas las PCs registradas"""
        if not self.ips_registradas:
            print("❌ No hay PCs registradas")
            return
        
        print("\n" + "="*50)
        print("📋 PCs REGISTRADAS")
        print("="*50)
        for ip, info in self.ips_registradas.items():
            print(f"IP: {ip} | Nombre: {info['nombre']} | Estado: {info['estado']}")
    
    def ver_datos_pc(self):
        """Muestra datos de una PC específica"""
        self.ver_todas_pcs()
        ip = input("\nIngresa la IP de la PC: ").strip()
        
        if ip not in self.ips_registradas:
            print("❌ IP no registrada")
            return
        
        print(f"\n⏳ Conectando a {self.ips_registradas[ip]['nombre']}...")
        datos = self.obtener_datos(ip)
        
        if datos and "error" not in datos:
            self.mostrar_datos_formateados(datos)
        else:
            print(f"❌ {datos.get('error', 'Error desconocido')}")
    
    def mostrar_datos_formateados(self, datos):
        """Muestra los datos de forma legible"""
        print("\n" + "="*60)
        print(f"🖥️  {datos.get('hostname', 'N/A')}")
        print("="*60)
        print(f"IP: {datos.get('ip', 'N/A')}")
        print(f"SO: {datos.get('sistema_operativo', 'N/A')} {datos.get('version_so', '')}")
        print(f"CPU: {datos.get('procesador', 'N/A')}")
        
        print(f"\n📊 CPU: {datos['cpu']['porcentaje']}%")
        print(f"   Núcleos: {datos['cpu']['nucleos_logicos']} (Físicos: {datos['cpu']['nucleos_fisicos']})")
        
        print(f"\n💾 MEMORIA: {datos['memoria']['porcentaje']}%")
        print(f"   Usada: {datos['memoria']['usada_gb']:.2f} GB")
        print(f"   Total: {datos['memoria']['total_gb']:.2f} GB")
        
        print(f"\n💿 DISCO:")
        for disco, info in datos['disco'].items():
            print(f"   {disco}: {info['porcentaje']}% ({info['usado_gb']:.2f}GB / {info['total_gb']:.2f}GB)")
        
        print(f"\n⚙️  PROCESOS: {datos['procesos']['total']} activos")
        print("   Top 5 por CPU:")
        for proc in datos['procesos']['procesos_activos'][:5]:
            print(f"   - {proc['nombre']}: CPU {proc['cpu_percent']}%, RAM {proc['memoria_mb']:.2f}MB")
        
        print(f"\n🌐 CONEXIONES: {datos['red']['conexiones_activas']} activas")
        print("="*60)
    
    def ejecutar_comando_remoto(self):
        """Ejecuta un comando en una PC remota"""
        self.ver_todas_pcs()
        ip = input("\nIngresa la IP: ").strip()
        
        if ip not in self.ips_registradas:
            print("❌ IP no registrada")
            return
        
        comando = input("Ingresa el comando (dir, tasklist, ipconfig, etc.): ").strip()
        
        print(f"⏳ Ejecutando comando...")
        resultado = self.ejecutar_comando(ip, comando)
        
        if "error" not in resultado:
            print(f"\n📤 Salida:")
            print(resultado.get('salida', ''))
            if resultado.get('error'):
                print(f"⚠️  {resultado['error']}")
        else:
            print(f"❌ {resultado['error']}")
    
    def verificar_estado_todas(self):
        """Verifica el estado de todas las PCs"""
        print("\n⏳ Verificando estado de todas las PCs...")
        
        for ip, info in self.ips_registradas.items():
            if self.verificar_conexion(ip):
                self.ips_registradas[ip]["estado"] = "✅ En línea"
                print(f"✅ {info['nombre']} ({ip})")
            else:
                self.ips_registradas[ip]["estado"] = "❌ Desconectada"
                print(f"❌ {info['nombre']} ({ip})")
        
        self.guardar_ips()
    
    def exportar_excel(self):
        """Exporta datos a Excel (requiere openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            # Encabezados
            headers = ["IP", "Nombre PC", "Estado", "CPU %", "Memoria %", "S.O.", "Timestamp"]
            ws.append(headers)
            
            # Aplicar estilo a encabezados
            fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
            
            # Datos
            for ip in self.ips_registradas:
                datos = self.obtener_datos(ip)
                if datos and "error" not in datos:
                    fila = [
                        ip,
                        self.ips_registradas[ip]['nombre'],
                        self.ips_registradas[ip]['estado'],
                        f"{datos['cpu']['porcentaje']}%",
                        f"{datos['memoria']['porcentaje']}%",
                        datos.get('sistema_operativo', 'N/A'),
                        datos.get('timestamp', 'N/A')
                    ]
                    ws.append(fila)
            
            archivo = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(archivo)
            print(f"✅ Exportado a {archivo}")
            
        except ImportError:
            print("❌ Instala openpyxl: pip install openpyxl")

if __name__ == "__main__":
    cliente = ClienteCentral()
    cliente.mostrar_menu()
