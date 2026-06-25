import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger("ExcelLogic")

def build_excel_workbook(db, target_ip: str | None = None) -> Workbook:
    """
    Construye un libro de Excel 'Lindo y Ordenado' extrayendo todas las métricas.
    """
    from database import PC, Event, Metric
    wb = Workbook()

    # ── Hoja 1: Resumen de Flota (PCs) ──
    ws = wb.active
    ws.title = "Resumen de Flota"

    headers = [
        "ID", "IP", "Nombre", "Hostname", "SO", "Estado",
        "Arquitectura", "Procesador",
        "CPU %", "RAM Total (GB)", "RAM Uso (GB)", "RAM %",
        "Disco %", "Temp Máx (ºC)", "Batería %", "Conexiones", "Procesos",
        "Último visto", "Registrado"
    ]
    ws.append(headers)

    # Codere Branding Headers
    h_fill = PatternFill(start_color="7EBB28", end_color="7EBB28", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    query = db.query(PC)
    if target_ip and str(target_ip).strip():
        query = query.filter(PC.ip == str(target_ip).strip())

    for pc in query.all():
        arch, proc, cpu, ram_tot, ram_use, ram_pct, disk_pct, temp, battery, conns, procs = ("", "", "", "", "", "", "", "", "", "", "")
        if pc.last_metrics:
            try:
                m = json.loads(pc.last_metrics)

                arch = m.get('system', {}).get('architecture', '')
                proc = m.get('system', {}).get('processor', '')

                cpu = f"{m.get('cpu', {}).get('percent', '')}%"
                ram_tot = m.get('memory', {}).get('total_gb', '')
                ram_use = m.get('memory', {}).get('used_gb', '')
                ram_pct = f"{m.get('memory', {}).get('percent', '')}%"

                disks = m.get('disk', {})
                if disks:
                    first_disk = list(disks.values())[0]
                    disk_pct = f"{first_disk.get('percent', '')}%"

                max_temp = 0.0
                temps_dict = m.get('temperatures', {})
                if temps_dict:
                    for sensor_list in temps_dict.values():
                        for sensor in sensor_list:
                            if sensor.get('current', 0) > max_temp:
                                max_temp = sensor.get('current', 0)
                if max_temp > 0:
                    temp = f"{round(max_temp, 1)} °C"

                bat = m.get('battery')
                if bat:
                    battery = f"{bat.get('percent', '')}%"

                conns = m.get('network', {}).get('connections', '')
                procs = m.get('processes', {}).get('total', '')
            except json.JSONDecodeError as e:
                logger.error("JSON corrupto en métricas para PC ID %s: %s", pc.id, e)
            except Exception as e:
                logger.error("Fallo imprevisto parseando métricas para Excel (PC %s): %s", pc.id, e, exc_info=True)

        ws.append([
            pc.id, pc.ip, pc.name, pc.hostname or "",
            pc.os or "", pc.status,
            arch, proc,
            cpu, ram_tot, ram_use, ram_pct, disk_pct, temp, battery, conns, procs,
            pc.last_seen or "", pc.registered_at
        ])

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Hoja 2: Eventos ──
    ws2 = wb.create_sheet("Historial Eventos")
    ws2.append(["PC", "IP", "Evento", "Timestamp", "Downtime (seg)"])
    for cell in ws2[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center")

    query_events = db.query(Event).order_by(Event.timestamp.desc()).limit(5000)
    if target_ip and str(target_ip).strip():
        query_events = query_events.filter(Event.pc_ip == str(target_ip).strip())

    for e in query_events.all():
        ws2.append([e.pc_name, e.pc_ip, e.type, e.timestamp, e.downtime_seconds or ""])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 25

    # ── Hoja 3: Métricas Crudas ──
    ws3 = wb.create_sheet("Métricas Históricas")
    ws3.append(["PC_ID", "Timestamp", "CPU %", "RAM %", "Disco %", "Procesos", "Conexiones"])
    for cell in ws3[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center")

    query_metrics = db.query(Metric)
    if target_ip and str(target_ip).strip():
        pc = db.query(PC).filter(PC.ip == str(target_ip).strip()).first()
        if pc:
            query_metrics = query_metrics.filter(Metric.pc_id == pc.id)

    query_metrics = query_metrics.order_by(Metric.timestamp.desc()).limit(10000)

    for m in query_metrics.all():
        ws3.append([m.pc_id, m.timestamp, m.cpu_percent, m.ram_percent,
                    m.disk_percent, m.processes_count, m.network_connections])

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    return wb
